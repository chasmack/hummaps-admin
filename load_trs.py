import psycopg2
from openpyxl import load_workbook
import time
import re

from create_funcs import subsec_bits
from create_funcs import township_number, range_number, township_str, range_str

from const import *

# load_trs.py - create and load the trs and source tables

def load_trs():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Create and load the source table
        print('CREATE TABLE: {table_source} ...'.format(table_source=TABLE_SOURCE))

        cur.execute("""
            DROP TABLE IF EXISTS {table_source} CASCADE;
            CREATE TABLE {table_source}  (
               id integer PRIMARY KEY,
               description text,
               quality integer
            );
        """.format(table_source=TABLE_SOURCE))
        con.commit()

        cur.execute("""
            INSERT INTO {table_source} (id, description, quality)
            VALUES
                ({trs_source_hollins_section}, 'Hollins full-section records', 10),
                ({trs_source_hollins_subsection}, 'Hollins subsection records', 40),
                ({trs_source_parsed_section}, 'Parsed full-section records', 40),
                ({trs_source_parsed_subsection}, 'Parsed subsection records', 50),
                ({trs_source_xlsx_data}, 'Additional XLSX records', 40)
            ;
        """.format(
            table_source=TABLE_SOURCE,
            trs_source_hollins_section=TRS_SOURCE_HOLLINS_SECTION,
            trs_source_hollins_subsection=TRS_SOURCE_HOLLINS_SUBSECTION,
            trs_source_parsed_section=TRS_SOURCE_PARSED_SECTION,
            trs_source_parsed_subsection=TRS_SOURCE_PARSED_SUBSECTION,
            trs_source_xlsx_data=TRS_SOURCE_XLSX_DATA
        ))
        con.commit()

        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Load the trs table from Hollins full section trs data
        print('CREATE TABLE: {table_trs} ...'.format(table_trs=TABLE_TRS))

        cur.execute("""
            DROP TABLE IF EXISTS {table_trs};
            CREATE TABLE {table_trs}  (
              id serial PRIMARY KEY,
              map_id integer REFERENCES {table_map},
              tshp integer,
              rng integer,
              sec integer,
              subsec integer,
              source_id integer REFERENCES {table_source}
            );
        """.format(
            table_trs=TABLE_TRS,
            table_map=TABLE_MAP,
            table_source=TABLE_SOURCE
        ))
        con.commit()

        # Insert full section trs records from Hollins trs data
        cur.execute("""
            WITH q1 AS (
            SELECT map_id,
                {function_township_number}(township) tshp,
                {function_range_number}(range) rng,
                regexp_split_to_table(trim(',' from section), ',')::int sec
            FROM {table_hollins_trs}
            ORDER BY map_id, tshp, rng, sec
            )
            INSERT INTO {table_trs} (map_id, tshp, rng, sec, source_id)
            SELECT q1.*, {trs_source_hollins_section}::integer source_id
            FROM q1
            ;
        """.format(
            table_trs=TABLE_TRS,
            table_hollins_trs=TABLE_HOLLINS_TRS,
            function_township_number=FUNCTION_TOWNSHIP_NUMBER,
            function_range_number=FUNCTION_RANGE_NUMBER,
            trs_source_hollins_section=TRS_SOURCE_HOLLINS_SECTION
        ))
        con.commit()

        print('INSERT (HOLLINS SECTION): ' + str(cur.rowcount) + ' rows effected.')

        # Insert trs records from Hollins subsection data
        cur.execute("""
            INSERT INTO {table_trs} (map_id, tshp, rng, sec, subsec, source_id)
            SELECT
                map_id, tshp, rng, sec, subsec,
                {trs_source_hollins_subsection} source_id
            FROM {function_hollins_subsec}()
            ;
        """.format(
            table_trs=TABLE_TRS,
            function_hollins_subsec=FUNCTION_HOLLINS_SUBSEC,
            trs_source_hollins_subsection=TRS_SOURCE_HOLLINS_SUBSECTION
        ))
        con.commit()

        print('INSERT (HOLLINS SUBSECTION): ' + str(cur.rowcount) + ' rows effected.')

        # Read additional map data from the XLSX file
        ws = load_workbook(filename=XLSX_DATA_MAP, read_only=True).active
        HEADER_LINES = 1

        maps = []
        for map in ws.iter_rows(min_row=HEADER_LINES + 1):
            maps.append(tuple(c.value for c in map))

        cur.executemany("""
            WITH q1 AS (
            SELECT
                (%s)::text maptype,
                (%s)::integer book,
                (%s)::integer page,
                (%s)::integer npages,
                (%s)::date recdate,
                (%s)::text surveyor,
                (%s)::text client,
                (%s)::text description,
                regexp_matches(
                  regexp_split_to_table((%s)::text, ',\s*'),
                  '(.* )?S(\d+) T?(\d+[NS]) R?(\d+[EW])'
                ) trs,
                (%s)::text note
            )
            INSERT INTO {table_trs} (
               map_id, tshp, rng, sec, subsec, source_id
            )
            SELECT m.id map_id,
                {function_township_number}(trs[3]) tshp,
                {function_range_number}(trs[4]) rng,
                trs[2]::integer sec,
                {function_subsec_bits}(trs[1]) subsec,
                {trs_source_xlsx_data}::integer soutce_id
            FROM q1
            JOIN {table_maptype} t ON q1.maptype = t.maptype
            JOIN {table_map} m ON t.id = m.maptype_id
                AND q1.book = m.book AND q1.page = m.page
            ;
        """.format(
            table_trs=TABLE_TRS,
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE,
            function_township_number=FUNCTION_TOWNSHIP_NUMBER,
            function_range_number=FUNCTION_RANGE_NUMBER,
            function_subsec_bits=FUNCTION_SUBSEC_BITS,
            trs_source_xlsx_data=TRS_SOURCE_XLSX_DATA
        ), maps)
        con.commit()

        print('INSERT (EXTRAS): ' + str(cur.rowcount) + ' rows effected.')

        # Vacuum must run outside of a transaction
        con.autocommit = True
        for t in (TABLE_TRS, TABLE_SOURCE):
            cur.execute('VACUUM FREEZE ' + t)


# load_trs_parsed_subsection.py
#
# Load additional trs subsection records parsed from map descriptions.
# Map descriptions containing subdivisional location information might look like -
#
# NE/4 S6 & SE/4 S6 1N,2E + SE/4 S31 2N,2E  HWY 36 & VAN DUZEN RIVER
#
# The initial query parses the map description picking out subdivisional sections
# and township/range qualifiers. The records passed to subsequent processing looks like -
#
#   id   |    subsec     |                              trs
# -------+---------------+-----------------------------------------------------------------------
#   5547 | NE/4 S6       | {"S7 1N 2E","S5 1N 2E","S6 1N 2E","S32 2N 2E","S31 2N 2E","S8 1N 2E"}
#   5547 | SE/4 S6       | {"S7 1N 2E","S5 1N 2E","S6 1N 2E","S32 2N 2E","S31 2N 2E","S8 1N 2E"}
#   5547 | 1N,2E         | {"S7 1N 2E","S5 1N 2E","S6 1N 2E","S32 2N 2E","S31 2N 2E","S8 1N 2E"}
#   5547 | SE/4 S31      | {"S7 1N 2E","S5 1N 2E","S6 1N 2E","S32 2N 2E","S31 2N 2E","S8 1N 2E"}
#   5547 | 2N,2E         | {"S7 1N 2E","S5 1N 2E","S6 1N 2E","S32 2N 2E","S31 2N 2E","S8 1N 2E"}
#
# The township/range is often missing from the description. In that case the information
# from existing non-subsection records is used to determine the township and range.
# If township/range information is present in the description it is compared with
# the non-subsection trs data to validate the township/range/section.

def load_trs_parsed_subsection():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Patterns used with matches to break out sections with a subsection or a township/range.
        SUBSEC_PATTERN = '(?:(?:[NS][EW]/4|[NSEW]/2)\s*)?(?:[NS][EW]/4|[NSEW]/2)\s*S\d{1,2}'
        TR_PATTERN = '\d+[NS]\s*,\s*\d+[EW]'

        # Get maps with no subsection data and a subsection in the description.
        cur.execute("""
            WITH q1 AS (
              -- find maps with no subsection information
              -- aggregate township/range/section records
              SELECT m.id,
                array_agg(concat_ws(' ',
                  'S' || trs.sec::text,
                  {function_township_str}(trs.tshp),
                  {function_range_str}(trs.rng)
                )) trs
              FROM {table_map} m
              JOIN {table_trs} trs ON trs.map_id = m.id
              GROUP BY m.id
              HAVING bit_or(subsec) IS NULL
            )
            SELECT m.id map_id, q1.trs,
              unnest(regexp_matches(m.description, '{subsec_pattern}|{tr_pattern}', 'ig')) subsec
            FROM q1
            JOIN {table_map} m USING (id)
            WHERE m.description ~* '{subsec_pattern}'
            ;
        """.format(
            table_trs=TABLE_TRS,
            table_map=TABLE_MAP,
            function_township_str=FUNCTION_TOWNSHIP_STR,
            function_range_str=FUNCTION_RANGE_STR,
            subsec_pattern=SUBSEC_PATTERN,
            tr_pattern=TR_PATTERN
        ))
        con.commit()

        # Combine records into maps with a list of subsections.
        # If a township/range qualifier is found it is placed at the end
        # of the subsection list and a new map list started. Only some lists
        # will have a township/range qualifier as their last element.
        maps = []
        for map_id, trs, subsec in cur:
            if len(maps) == 0 or maps[-1][0] != map_id or re.fullmatch(TR_PATTERN, maps[-1][-1][-1], re.IGNORECASE):
                maps.append((map_id, trs, [subsec]))
            else:
                maps[-1][-1].append(subsec)

        # Convert each subsection in the subsection list into a trs record.
        recs = []
        for map_id, trs, subsecs in maps:

            # Check if last element of the subsection list is a township/raange qualifier.
            m = re.fullmatch('(\d+[NS]).*(\d+[EW])', subsecs[-1])
            if m:
                tshp, rng = m.groups()
                subsecs.pop()
            else:
                tshp = rng = None

            for ss in subsecs:

                # Break out section and subsection information
                m = re.fullmatch('(.*[24])\s*(S\d+)', ss)
                if m is None:
                    print('>> PARSED SUBSEC: Bad section. id=%d subsec=%s' % (map_id, ss))
                    continue
                subsec, sec = m.groups()

                # Make sure the subsection portion converts correctly.
                if subsec_bits(subsec) is None:
                    print('>> PARSED SUBSEC: Bad subsec: id=%d subsec=%s' % (map_id, ss))
                    continue

                if tshp is None or rng is None:

                    # Get township and range from trs data
                    n = 0
                    for s, t, r in (d.split() for d in trs):
                        if s == sec:
                            tshp, rng = t, r
                            n += 1
                    if n == 0:
                        print('>> PARSED SUBSEC: No trs record found. id=%d subsec=%s' % (map_id, ss))
                        continue
                    elif n > 1:
                        print('>> PARSED SUBSEC: Multiple trs records found. id=%d subsec=%s' % (map_id, ss))
                        continue

                else:
                    # Verify township/range/section is in the trs data.
                    if '%s %s %s' % (sec, tshp, rng) not in trs:
                        print('>> PARSED SUBSEC: Township/range/section not in trs. id=%d subsec=%s' % (map_id, ss))
                        continue

                recs.append((map_id, township_number(tshp), range_number(rng), int(sec[1:]), subsec_bits(subsec)))

        cur.executemany("""
            INSERT INTO {table_trs} (map_id, tshp, rng, sec, subsec, source_id)
            VALUES (%s, %s, %s, %s, %s, {trs_source_parsed_subsection});
        """.format(
            table_trs=TABLE_TRS,
            trs_source_parsed_subsection=TRS_SOURCE_PARSED_SUBSECTION
        ), recs)
        con.commit()

        print('INSERT (PARSED SUBSEC): %d rows affected.' % cur.rowcount)

        # Vacuum must run outside of a transaction
        con.autocommit = True
        cur.execute('VACUUM FREEZE ' + TABLE_TRS)


if __name__ == '__main__':

    print('\nCreating staging tables ... ')
    startTime = time.time()

    load_trs()
    load_trs_parsed_subsection()

    endTime = time.time()
    print('{0:.3f} sec'.format(endTime - startTime))
