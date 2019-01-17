import psycopg2
from openpyxl import load_workbook
from datetime import datetime
import xml.etree.ElementTree as etree
import xml.dom.minidom as minidom

from hollins.const import *

#
# export_xml - Create an XML file from Hummaps data for import into Hollins tables.mdb
#

def export_xml(xml_file):

    time = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    attrib = {
        'generated': time,
        'xmlns:od': 'urn:schemas-microsoft-com:officedata'
    }
    root = etree.Element('dataroot', attrib=attrib)

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Extract the SURVEYOR lastnames and create surveyor full names.
        cur.execute("""
            SELECT id surveyor_id, lastname,
                firstname ||
                coalesce(' ' || left(secondname, 1), '') ||
                coalesce(' ' || left(thirdname, 1), '') ||
                ' ' || lastname ||
                coalesce(' ' || suffix, '') AS surveyor
            FROM {table_surveyor}
            ORDER BY lastname, firstname;
        """.format(
            table_surveyor=TABLE_PROD_SURVEYOR)
        )

        nrecs = 0

        # Add a record for unknown surveyors
        elem = etree.SubElement(root, 'Surveyor')
        etree.SubElement(elem, 'Surveyor').text = 'UNKNOWN'
        etree.SubElement(elem, 'lastname').text = 'UNKNOWN'
        nrecs += 1

        # A list of surveyors indexed by full name
        surveyor_list = {}

        for surveyor_id, lastname, surveyor in cur:
            # Surveyor full name must be unique
            if surveyor in surveyor_list:
                print('Error: Duplicate surveyor name: %s' % surveyor)
                exit(1)
            surveyor_list[surveyor] = [surveyor_id, lastname]

            elem = etree.SubElement(root, 'Surveyor')
            etree.SubElement(elem, 'Surveyor').text = surveyor
            etree.SubElement(elem, 'lastname').text = lastname
            nrecs += 1

        # Reindex the surveyor list by surveyor_id
        temp = surveyor_list
        surveyor_list = {}
        for surveyor in temp.keys():
            surveyor_id, lastname = temp[surveyor]
            surveyor_list[surveyor_id] = [lastname, surveyor]
        del temp

        print('Created %d SURVEYOR records.' % nrecs)

        # Extract the MAP records.
        cur.execute("""
            WITH q1 AS (
                SELECT m.id map_id,
                    array_remove(array_agg(s.id), NULL) as surveyor_ids
                FROM {table_map} m
                LEFT JOIN {table_signed_by} sb ON sb.map_id = m.id
                LEFT JOIN {table_surveyor} s ON sb.surveyor_id = s.id
                GROUP BY m.id
            )
            SELECT m.id map_id, t.maptype, m.book, m.page firstpage, m.page + m.npages - 1 lastpage,
              m.recdate, q1.surveyor_ids, m.client donefor, m.description descrip, pdf.pdffile
            FROM {table_map} m
            JOIN {table_maptype} t ON t.id = m.maptype_id
            LEFT JOIN q1 ON q1.map_id = m.id
            LEFT JOIN {table_pdf} ON pdf.map_id = m.id
            ORDER BY map_id;
        """.format(
            table_map=TABLE_PROD_MAP,
            table_maptype=TABLE_PROD_MAPTYPE,
            table_signed_by=TABLE_PROD_SIGNED_BY,
            table_surveyor=TABLE_PROD_SURVEYOR,
            table_pdf=TABLE_PROD_PDF)
        )

        # A list of MAP elements indexed by map_id
        map_records = {}

        for row in cur:
            map_id, maptype, book, firstpage, lastpage, recdate, surveyor_ids, donefor, descrip, pdffile = row

            elem = etree.Element('map')
            etree.SubElement(elem, 'ID').text = str(map_id)
            etree.SubElement(elem, 'maptype').text = maptype
            etree.SubElement(elem, 'BOOK').text = str(book)
            etree.SubElement(elem, 'FIRSTPAGE').text = str(firstpage)
            etree.SubElement(elem, 'LASTPAGE').text = str(lastpage)

            if recdate is not None:
                etree.SubElement(elem, 'RECDATE').text = recdate.strftime('%Y-%m-%dT00:00:00')

            if len(surveyor_ids) == 0:
                surveyors = 'UNKNOWN'
            else:
                surveyors = sorted([surveyor_list[id] for id in surveyor_ids], key=lambda s: ' '.join(s))
                surveyors = ' & '.join(s[1] for s in surveyors)
            etree.SubElement(elem, 'SURVEYOR').text = surveyors

            if donefor is not None:
                etree.SubElement(elem, 'DONEFOR').text = donefor

            if descrip is not None:
                etree.SubElement(elem, 'DESCRIP').text = descrip

            if pdffile is not None:
                etree.SubElement(elem, 'Picture').text = pdffile

            map_records[map_id] = elem

        # Relate Hummaps subsection bit positions (lsb -> msb) to Hollins subsection numbers
        hollins_subsecs = (4, 3, 2, 1, 5, 6, 7, 8, 12, 11, 10, 9, 13, 14, 15, 16)

        # Read the subsection list
        ws = load_workbook(filename=XLSX_SUBSECTION_LIST, read_only=True).active
        for trs_subsec in (row[0].value for row in ws.iter_rows()):

            # Add subsection data for one section at a time.
            cur.execute("""
                SELECT m.id map_id, trs.subsec
                FROM {table_map} m
                JOIN {table_trs} trs ON trs.map_id = m.id
                WHERE trs.source_id = 1
                AND trs.tshp = {function_township_number}('{township}')
                AND trs.rng = {function_range_number}('{range}')
                AND trs.sec = {section};
            """.format(
                table_map=TABLE_PROD_MAP,
                table_trs=TABLE_PROD_TRS,
                function_township_number=FUNCTION_TOWNSHIP_NUMBER,
                function_range_number=FUNCTION_RANGE_NUMBER,
                township=trs_subsec[0:3],
                range=trs_subsec[3:5],
                section=int(trs_subsec[5:7]))
            )

            for map_id, subsec_bits in cur:
                subsec_list = []
                for i in range(16):
                    if (1 << i) & subsec_bits > 0:
                        subsec_list.append(hollins_subsecs[i])
                subsec_list = ','.join((str(n) for n in sorted(subsec_list)))
                elem = map_records[map_id]

                if trs_subsec[0].isdigit():
                    # Can't start an element label with a digit.
                    trs_subsec = '_x%04x_%s' % (ord(trs_subsec[0]), trs_subsec[1:])

                etree.SubElement(elem, trs_subsec).text = ',%s,' % subsec_list

        nrecs = 0
        for id in sorted(map_records.keys()):
            root.append(map_records[id])
            nrecs += 1

        print('Created %d MAP records.' % nrecs)

        # Extract the TRS records.
        cur.execute("""
            SELECT map_id,
              lpad({function_township_str}(tshp), 3, '0') tshp,
              {function_range_str}(rng) rng,
              ',' || string_agg(sec::text, ',') || ',' secs
            FROM (
              SELECT DISTINCT map_id, tshp, rng, sec
              FROM {table_trs}
              WHERE source_id = 0
              ORDER BY map_id, tshp, rng, sec
            ) q1
            GROUP BY map_id, tshp, rng
            ORDER BY map_id;
        """.format(
            table_trs=TABLE_PROD_TRS,
            function_township_str=FUNCTION_TOWNSHIP_STR,
            function_range_str=FUNCTION_RANGE_STR)
        )

        nrecs = 0
        for row in cur:
            map_id, tshp, rng, secs = row
            elem = etree.SubElement(root, 'TRS')
            etree.SubElement(elem, 'ID').text = str(map_id)
            # Unknown township/range appear as None
            etree.SubElement(elem, 'TOWNSHIP').text = tshp if tshp else '0'
            etree.SubElement(elem, 'RANGE').text = rng if rng else '0'
            etree.SubElement(elem, 'SECTION').text = secs
            nrecs += 1

        print('Created %d TRS records.' % nrecs)

    # with open(xml_file, 'wb') as f:
    #     etree.ElementTree(root).write(f)

    # Reparse the etree xml with minidom and write pretty xml.
    dom = minidom.parseString(etree.tostring(root, encoding='utf-8'))
    with open(xml_file, 'wb') as f:
        f.write(dom.toprettyxml(indent='  ', encoding='utf-8'))

    return


def check_trs(xml_file, check_file):

    trs = {}
    nrecs = 0
    root = etree.parse(xml_file).getroot()
    for elem in root.findall('TRS'):
        map_id = elem.find('ID').text
        tshp = elem.find('TOWNSHIP').text
        rng = elem.find('RANGE').text
        secs = elem.find('SECTION').text
        key = '_'.join((map_id, tshp, rng))
        trs[key] = secs
        nrecs += 1

    print('Found %d records in %s.' % (nrecs, xml_file))

    nrecs = 0
    root = etree.parse(check_file).getroot()
    for elem in root.findall('TRS'):
        map_id = elem.find('ID').text
        tshp = elem.find('TOWNSHIP').text
        rng = elem.find('RANGE').text
        secs = elem.find('SECTION').text
        key = '_'.join((map_id, tshp, rng))
        if key not in trs:
            print('Missing record: map_id=%s tshp=%s rng=%s' % (map_id, tshp, rng))
        elif trs[key] != secs:
            print('Bad record: map_id=%s tshp=%s rng=%s expected=%s found=%s' % (map_id, tshp, rng, trs[key], secs))
        trs['_'.join((map_id, tshp, rng))]
        nrecs += 1

    print('Compared %d records in %s.' % (nrecs, check_file))


if __name__ == '__main__':

    export_xml('hollins/hollins.xml')
    check_trs('hollins/hollins.xml', 'hollins/update64/update64_trs.xml')
