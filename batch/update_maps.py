import psycopg2
import requests
from openpyxl import Workbook, load_workbook
import time
import re

XLSX_FILE = r'd:\Projects\Python\hummaps-admin\batch\test.xlsx'

PG_HOST = 'p3'
PG_USER_ADMIN = 'pgadmin'
PG_DATABASE_PROD = 'production'

# IPV4 connection
DSN_PROD = 'dbname={database} user={user} host={host}'.format(
    database=PG_DATABASE_PROD,
    user=PG_USER_ADMIN,
    host=PG_HOST
)

SCHEMA_STAGING = 'hummaps_staging'

TABLE_STAGING_HOLLINS_MAP = SCHEMA_STAGING + '.' + 'hollins_map'
TABLE_STAGING_HOLLINS_MAP_QQ = SCHEMA_STAGING + '.' + 'hollins_map_qq'
TABLE_STAGING_HOLLINS_SUBSECTION_LIST = SCHEMA_STAGING + '.' + 'hollins_subsection_list'
TABLE_STAGING_HOLLINS_SURVEYOR = SCHEMA_STAGING + '.' + 'hollins_surveyor'
TABLE_STAGING_HOLLINS_TRS = SCHEMA_STAGING + '.' + 'hollins_trs'
TABLE_STAGING_HOLLINS_FULLNAME = SCHEMA_STAGING + '.' + 'hollins_fullname'

TABLE_STAGING_CC = SCHEMA_STAGING + '.' + 'cc'
TABLE_STAGING_CC_IMAGE = SCHEMA_STAGING + '.' + 'cc_image'
TABLE_STAGING_MAP = SCHEMA_STAGING + '.' + 'map'
TABLE_STAGING_MAP_IMAGE = SCHEMA_STAGING + '.' + 'map_image'
TABLE_STAGING_MAPTYPE = SCHEMA_STAGING + '.' + 'maptype'
TABLE_STAGING_PDF = SCHEMA_STAGING + '.' + 'pdf'
TABLE_STAGING_SCAN = SCHEMA_STAGING + '.' + 'scan'
TABLE_STAGING_SIGNED_BY = SCHEMA_STAGING + '.' + 'signed_by'
TABLE_STAGING_SOURCE = SCHEMA_STAGING + '.' + 'source'
TABLE_STAGING_SUBSEC_NAMES = SCHEMA_STAGING + '.' + 'subsec_names'
TABLE_STAGING_SURVEYOR = SCHEMA_STAGING + '.' + 'surveyor'
TABLE_STAGING_TRS = SCHEMA_STAGING + '.' + 'trs'
TABLE_STAGING_TRS_PATH = SCHEMA_STAGING + '.' + 'trs_path'

SEQUENCE_MAP_ID = SCHEMA_STAGING + '.' + 'map_id_seq'

FUNCTION_MAP_ID = SCHEMA_STAGING + '.' + 'map_id'
FUNCTION_MAP_NAME = SCHEMA_STAGING + '.' + 'map_name'
FUNCTION_TOWNSHIP_NUMBER = SCHEMA_STAGING + '.' + 'township_number'
FUNCTION_TOWNSHIP_STR = SCHEMA_STAGING + '.' + 'township_str'
FUNCTION_RANGE_NUMBER = SCHEMA_STAGING + '.' + 'range_number'
FUNCTION_RANGE_STR = SCHEMA_STAGING + '.' + 'range_str'
FUNCTION_SUBSEC_BITS = SCHEMA_STAGING + '.' + 'subsec_bits'
FUNCTION_SUBSEC_STR = SCHEMA_STAGING + '.' + 'subsec_str'
FUNCTION_HOLLINS_SUBSEC = SCHEMA_STAGING + '.' + 'hollins_subsec'

SCHEMA_PROD = 'hummaps'

TABLE_PROD_CC = SCHEMA_PROD + '.' + 'cc'
TABLE_PROD_CC_IMAGE = SCHEMA_PROD + '.' + 'cc_image'
TABLE_PROD_MAP = SCHEMA_PROD + '.' + 'map'
TABLE_PROD_MAP_IMAGE = SCHEMA_PROD + '.' + 'map_image'
TABLE_PROD_MAPTYPE = SCHEMA_PROD + '.' + 'maptype'
TABLE_PROD_PDF = SCHEMA_PROD + '.' + 'pdf'
TABLE_PROD_SCAN = SCHEMA_PROD + '.' + 'scan'
TABLE_PROD_SIGNED_BY = SCHEMA_PROD + '.' + 'signed_by'
TABLE_PROD_SOURCE = SCHEMA_PROD + '.' + 'source'
TABLE_PROD_SURVEYOR = SCHEMA_PROD + '.' + 'surveyor'
TABLE_PROD_TRS = SCHEMA_PROD + '.' + 'trs'
TABLE_PROD_TRS_PATH = SCHEMA_PROD + '.' + 'trs_path'


def cleanup_surveyors():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:
        cur.execute("""
            SELECT
                initcap(concat_ws(' ', firstname, secondname, thirdname, lastname, suffix)) n1,
                concat_ws(' ', firstname, substring(secondname for 1), substring(thirdname for 1), lastname, suffix) n2
            FROM {table_surveyor};
        """.format(table_surveyor=TABLE_PROD_SURVEYOR))

        surveyor = dict(cur)

    wb = load_workbook(XLSX_FILE)
    ws = wb['Maps']
    ws_cols = list(c.value for c in ws[1])
    for cell in (dict(zip(ws_cols, row)) for row in ws.iter_rows(min_row=2)):

        key = re.sub('\s+\(.*', '', cell['SURVEYORS'].value)
        if key in surveyor:
            cell['SURVEYORS'].value = surveyor[key]
        elif key not in surveyor.values():
            print('Surveyor not found: %s' % cell['SURVEYORS'].value)
        else:
            print('Skipping: %s' % cell['SURVEYORS'].value)

    wb.save(filename=XLSX_FILE)


def update_maps():

    wb = load_workbook(XLSX_FILE)
    ws = wb['Maps']

    ws_cols = list(c.value for c in ws[1])

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        missing_map_images = []
        missing_pdfs = []

        for cell in (dict(zip(ws_cols, row)) for row in ws.iter_rows(min_row=2)):

            cur.execute("""
                SELECT m.id, t.maptype, lower(t.abbrev) abbrev, m.book, m.page,
                  count(distinct i.id) map_images, count(distinct pdf.id) pdfs
                FROM {table_map} m
                JOIN {table_maptype} t ON t.id = m.maptype_id
                LEFT JOIN {table_map_image} i ON i.map_id = m.id
                LEFT JOIN {table_pdf} pdf ON pdf.map_id = m.id
                WHERE t.maptype = %s AND m.book = %s AND m.page = %s
                GROUP BY m.id, t.maptype, t.abbrev, m.book, m.page
                ;
            """.format(
                table_map=TABLE_PROD_MAP,
                table_maptype=TABLE_PROD_MAPTYPE,
                table_map_image=TABLE_PROD_MAP_IMAGE,
                table_pdf=TABLE_PROD_PDF,
            ), (cell['MAPTYPE'].value, cell['BOOK'].value, cell['PAGE'].value))

            for row in cur:
                map_id = row[0]
                abbrev, book, page = row[2:5]
                map_images, pdfs = row[5:7]

                cell['MAP_ID'].value = map_id

                url_base = 'https://hummaps.com'
                map_name = '%03d%s%03d' % (book, abbrev, page)
                image_base = '/map/%s/%03d/%s' % (abbrev, book, map_name)
                imagefiles = []
                while True:
                    f = image_base + '-%03d.jpg' % (len(imagefiles) + 1)
                    r = requests.head(url_base + f)
                    if r.status_code == 200:
                        imagefiles.append(f)
                        print('%s: %d' % (f, r.status_code))
                    else:
                        break

                if map_images > 0 and map_images != len(imagefiles):
                    print('WARNING: %s: image_pages=%d imagefiles=%d' % (map_name, map_images, len(imagefiles)))
                else:
                    cell['MAP_IMAGES'].value = len(imagefiles)
                if map_images == 0:
                    for i in range(len(imagefiles)):
                        missing_map_images.append((map_id, i + 1, imagefiles[i]))

                pdffile = '/pdf/%s/%03d/%s.pdf' % (abbrev, book, map_name)
                r = requests.head(url_base + pdffile)
                if r.status_code == 200:
                    cell['PDFS'].value = 1
                    print('%s: %d' % (pdffile, r.status_code))
                    if pdfs == 0:
                        missing_pdfs.append((map_id, pdffile))

        cur.executemany("""
            INSERT INTO {table_map_image} (map_id, page, imagefile)
            VALUES (%s, %s, %s);
        """.format(
            table_map_image=TABLE_PROD_MAP_IMAGE,
        ), missing_map_images)
        con.commit()

        print('INSERT map_image: %d rows effected' % cur.rowcount)

        cur.executemany("""
            INSERT INTO {table_pdf} (map_id, pdffile)
            VALUES (%s, %s);
        """.format(
            table_pdf=TABLE_PROD_PDF,
        ), missing_pdfs)
        con.commit()

        print('INSERT pdf: %d rows effected' % cur.rowcount)

        wb.save(filename=XLSX_FILE)


if __name__ == '__main__':

    cleanup_surveyors()
