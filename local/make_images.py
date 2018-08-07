import os, shutil
import os.path
from glob import glob
from PIL import Image
from openpyxl import load_workbook
from subprocess import check_call
import time
import re

from local.const import *


# Extract single page tiff images from legacy multi-page map images

def extract_images():

    print('\nExtracting map images...')

    # Turn of warnings about oversize image files
    Image.warnings.simplefilter('ignore', Image.DecompressionBombWarning)

    nfiles = 0
    nframes = 0
    nccs = 0

    def sort_key(s):
        return re.sub('.*(\d{3})(\D{2})(\d{3}).*', '\\2\\1\\3', s)

    cc_dir = os.path.join(SCAN_DIR, 'cc')
    os.makedirs(cc_dir, exist_ok=True)

    for imagefile in sorted(glob(os.path.join(SOURCE_DIR, '*.tif')), key=sort_key):

        src_file = os.path.basename(imagefile).lower()
        book, maptype = re.match('(\d+)(\D+)', src_file).groups()

        map_dir = os.path.join(SCAN_DIR, maptype, book)
        os.makedirs(map_dir, exist_ok=True)

        print('\n' + src_file)

        with Image.open(imagefile) as img:

            # Process one frame at at time
            frame_number = 0
            last_frame = False
            while not last_frame:

                frame = img.copy()
                try:
                    frame_number += 1
                    img.seek(frame_number)

                except EOFError:
                    last_frame = True

                print('Frame %d' % (frame_number))
                print('Mode: %s' % ({
                    '1': '1-bit black and white',
                    'L': '8-bit greyscale',
                    'P': '8-bit color map',
                    'RGB': 'RGB color',
                    'RGBA': 'RGBa color'
                }[frame.mode]))

                # Calculate image size for the frame
                scan_dpi = tuple(int(round(d)) for d in frame.info['dpi'])

                # Default 96 dpi usually means dpi not set in the image header, assume 200 dpi
                if scan_dpi == (96, 96):
                    scan_dpi = (200, 200)
                    print('Scan dpi not set, using %s dpi' % (str(scan_dpi)))

                scan_size = tuple(d / dpi for d, dpi in zip(frame.size, scan_dpi))

                print(
                    'Scan size: %s @ %s dpi => %.2f x %.2f' % (str(frame.size), str(scan_dpi), scan_size[0], scan_size[1]))

                # Convert 8-bit color map to RGB
                if frame.mode == 'P':
                    print('Converting to RGB...')
                    frame = frame.convert('RGB')

                # Add the page number to the file name
                dest_file = '%s-%03d.tif' % (os.path.splitext(src_file)[0], frame_number)

                # If width of a recorded map (PM, RM, RS) is less than MIN_WIDTH assume frame is a CC
                MIN_WIDTH = 8.75
                if maptype.upper() in ('PM', 'RM', 'RS') and min(scan_size) < MIN_WIDTH:
                    dest = os.path.join(cc_dir, dest_file)
                    nccs += 1
                else:
                    dest = os.path.join(map_dir, dest_file)

                # Save the tiff scan
                frame.save(dest, resolution=float(scan_dpi[0]), resolution_unit='inch')

        nfiles += 1
        nframes += frame_number

    print('\n%d frames from %d files (%d CCs)' % (nframes, nfiles, nccs))


# Convert tiff map images in the scan directory to jpegs

def convert_maps():

    print('\nConverting map images...')

    # Turn of warnings about oversize image files
    Image.warnings.simplefilter('ignore', Image.DecompressionBombWarning)

    nfiles = 0

    for maptype in sorted(MAPTYPES.values()):

        for imagefile in sorted(glob(os.path.join(SCAN_DIR, maptype.lower(), '*',  '*.tif'))):

            src_path, src_file = os.path.split(imagefile.lower())
            dest_file = os.path.splitext(src_file)[0] + '.jpg'
            book = os.path.basename(src_path)

            dest_dir = os.path.join(MAP_DIR, maptype.lower(), book)
            os.makedirs(dest_dir, exist_ok=True)

            print('\n' + dest_file)

            with Image.open(imagefile) as img:

                print('Mode: %s' % ({
                    '1': '1-bit black and white',
                    'L': '8-bit greyscale',
                    'P': '8-bit color map',
                    'RGB': 'RGB color',
                    'RGBA': 'RGBa color'
                }[img.mode]))

                # Calculate the map image size
                scan_dpi = tuple(int(round(d)) for d in img.info['resolution'])
                map_dpi = tuple(min(MAP_DPI, dpi) for dpi in scan_dpi)
                map_size = tuple(int(round(d * mdpi / sdpi)) for d, mdpi, sdpi in zip(img.size, map_dpi, scan_dpi))

                print('Map size: %s @ %d dpi => %s @ %d dpi' % (str(img.size), scan_dpi[0], str(map_size), map_dpi[0]))

                # Convert 8-bit color map to RGB
                if img.mode == 'P':
                    print('Converting to RGB...')
                    img = img.convert('RGB')

                # Convert 1-bit black and white to 8-bit greyscale
                if img.mode == '1':
                    print('Converting to 8-bit greyscale...')
                    img = img.convert('L')

                # Resize and save the map image
                img = img.resize(map_size, resample=Image.BICUBIC)
                img.save(os.path.join(dest_dir, dest_file), dpi=map_dpi, quality=MAP_QUALITY)

                nfiles += 1

    print('\n%d files converted' % (nfiles))


# Convert tiff CC images in the scan directory to jpegs

def convert_ccs():

    print('\nConverting cc images...\n')

    # Turn of warnings about oversize image files
    Image.warnings.simplefilter('ignore', Image.DecompressionBombWarning)

    dest_dir = os.path.join(MAP_DIR, 'cc')
    os.makedirs(dest_dir, exist_ok=True)

    nfiles = 0

    for imagefile in sorted(glob(os.path.join(SCAN_DIR, 'cc', '*.tif'))):

        src_path, src_file = os.path.split(imagefile.lower())

        # Only convert properly named CCs
        if not re.fullmatch('\d{4}-(or|doc)-\d{6}-\d{3}.tif', src_file):
            print('Bad CC filename: %s' % (src_file))
            continue

        dest_file = os.path.splitext(src_file)[0] + '.jpg'

        print('\n' + dest_file)

        with Image.open(imagefile) as img:

            print('Mode: %s' % ({
                '1': '1-bit black and white',
                'L': '8-bit greyscale',
                'P': '8-bit color map',
                'RGB': 'RGB color',
                'RGBA': 'RGBa color'
            }[img.mode]))

            # Calculate the map image size
            scan_dpi = tuple(int(round(d)) for d in img.info['resolution'])
            map_dpi = tuple(min(MAP_DPI, dpi) for dpi in scan_dpi)
            map_size = tuple(int(round(d * mdpi / sdpi)) for d, mdpi, sdpi in zip(img.size, map_dpi, scan_dpi))

            print('Map size: %s @ %d dpi => %s @ %d dpi' % (str(img.size), scan_dpi[0], str(map_size), map_dpi[0]))

            # Convert 8-bit color map to RGB
            if img.mode == 'P':
                print('Converting to RGB...')
                img = img.convert('RGB')

            # Convert 1-bit black and white to 8-bit greyscale
            if img.mode == '1':
                print('Converting to 8-bit greyscale...')
                img = img.convert('L')

            # Resize and save the map image
            img = img.resize(map_size, resample=Image.BICUBIC)
            img.save(os.path.join(dest_dir, dest_file), dpi=map_dpi, quality=MAP_QUALITY)

            nfiles += 1

    print('\n%d files converted' % (nfiles))


# Combine jpeg map images and any jpeg CC images into pdfs

def make_pdfs():

    print('\nCreating PDFs...\n')

    # Read the CC data from the XLSX workbook
    ws = load_workbook(filename=XLSX_DATA_CC, read_only=True).active
    HEADER_LINES = 1

    cc_data = {}
    for row in ws.iter_rows(min_row=HEADER_LINES + 1):
        maptype, book, page, recdate, surveyor, doc_number, npages = list(c.value for c in row)

        if maptype in MAPTYPES:
            map = '%03d%s%03d' % (book, MAPTYPES[maptype].lower(), page)
        else:
            print('ERROR: Bad MAPTYPE in %s: %s' % (XLSX_DATA_CC, maptype))
            exit(1)

        m1 = re.fullmatch('(\d+) OR (\d+)', doc_number)
        m2 = re.fullmatch('(\d{4})-(\d+)-\d+', doc_number)
        if (m1):
            doc = '{0:04d}-or-{1:06d}'.format(*(int(n) for n in m1.groups()))
        elif (m2):
            doc = '{0:04d}-doc-{1:06d}'.format(*(int(n) for n in m2.groups()))
        else:
            print('ERROR: formant error in %s: "%s"' % (XLSX_DATA_CC, '" "'.join(row)))
            exit(1)

        cc_data[map] = cc_data.get(map, []) + [{'doc': doc, 'npages': npages}]

    # List of maps with missing CCs
    missing_ccs = []

    nfiles = 0
    for maptype in sorted(MAPTYPES.values()):

        # Get a list of map images indexed by map name
        mapimages = {}
        for imagefile in sorted(glob(os.path.join(MAP_DIR, maptype, '*', '*.jpg'))):
            map = re.match('^(\d{3}\D{2}\d{3})', os.path.basename(imagefile)).groups()[0]
            mapimages[map] = mapimages.get(map, []) + [imagefile]


        for map, imagefiles in sorted(mapimages.items(), key=lambda k: k[0]):

            # Append CC images to the end of the map image list
            if maptype.upper() in ('PM', 'RM', 'RS') and map in cc_data:
                for cc in sorted(cc_data[map], key=lambda cc: cc['doc'], reverse=True):
                    ccimages = sorted(glob(os.path.join(MAP_DIR, 'cc', cc['doc'] + '-*.jpg')))
                    if len(ccimages) == 0:
                        print('WARNING: Missing CC for %s: %s' % (map, cc['doc']))
                        missing_ccs.append(cc['doc'])
                    elif len(ccimages) != cc['npages']:
                        print('ERROR: Bad CC page count for %s: %s' % (map, cc['doc']))
                        exit(1)
                    else:
                        mapimages[map] += ccimages

            book = map[0:3]
            dest = os.path.join(PDF_DIR, maptype.lower(), book, map + '.pdf')
            os.makedirs(os.path.dirname(dest), exist_ok=True)

            magick_cmd = [MAGICK] + imagefiles + [dest]

            print('%s: %s' % (map, dest))
            check_call(magick_cmd)
            nfiles += 1

    print('\n%d PDFs created.' % nfiles)
    if (missing_ccs):
        print('Missing CCs: %s' % ', '.join(missing_ccs))


if __name__ == '__main__':

    print('\nMaking map images ... ')
    startTime = time.time()

    extract_images()
    convert_maps()
    convert_ccs()
    make_pdfs()

    endTime = time.time()
    print('\n{0:.3f} sec'.format(endTime - startTime))
