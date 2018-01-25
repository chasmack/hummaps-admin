from openpyxl import load_workbook
import psycopg2
import time

from const import *

# load_surveyor.py - create and load the surveyor and signed_by tables from XLSX data

# The file surveyor.xlsx contains data for the production surveyor table.
# The column hollins_fullname provides a mapping from hollins_map.surveyor to the
# production surveyor table. Historical errors have resulted in cases where many
# hollins surveyors map to a single production surveyor. This is handled by having
# multiple lines in surveyor.xlsx for a single production surveyor. In such cases
# all fields except hollins_fullname must be identical.

def load_surveyor():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Read the surveyor data and hollins surveyor mapping from the XLSX data
        ws = load_workbook(filename=XLSX_DATA_SURVEYOR, read_only=True).active
        HEADER_LINES = 1

        surveyor = {}
        hollins_mapping = {}
        for row in ws.iter_rows(min_row=HEADER_LINES + 1):
            hollins_fullname = row[0].value
            surveyor_fullname = row[1].value

            # check for duplicates and add hollins fullname to the mapping
            if hollins_fullname in hollins_mapping:
                print('Error processing {}: duplicate entries for hollins_fullname "{}"'.format(
                    XLSX_DATA_SURVEYOR, hollins_fullname))
                exit(-1)
            hollins_mapping[hollins_fullname] = surveyor_fullname

            if surveyor_fullname not in surveyor:
                # add a new surveyor
                surveyor[surveyor_fullname] = list(c.value for c in row[1:])

            elif surveyor[surveyor_fullname] != list(c.value for c in row[1:]):
                # fields are not identical for duplicate surveyor entries
                print('Error processing {}: duplicate entries not identical "{}"'.format(
                    XLSX_DATA_SURVEYOR, surveyor_fullname))
                exit(-1)

        # Create and populate the surveyor table
        print('CREATE TABLE: {table_surveyor} ...'.format(table_surveyor=TABLE_SURVEYOR))

        cur.execute("""
            DROP TABLE IF EXISTS {table_surveyor} CASCADE;
            CREATE TABLE {table_surveyor}  (
              id serial PRIMARY KEY,
              fullname text NOT NULL UNIQUE,
              firstname text,
              secondname text,
              thirdname text,
              lastname text,
              suffix text,
              pls text,
              rce text
            );
        """.format(table_surveyor=TABLE_SURVEYOR))

        cur.executemany("""
            INSERT INTO {table_surveyor} (
              fullname, firstname, secondname, thirdname, lastname, suffix, pls, rce
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
        """.format(table_surveyor=TABLE_SURVEYOR), list(surveyor.values()))
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))

        # Check the surveyor fullnames.
        cur.execute("""
            SELECT fullname, firstname, secondname, thirdname, lastname, suffix
            FROM {table_surveyor}
            WHERE fullname != concat_ws(' ', firstname, secondname, thirdname, lastname, suffix)
            ;
        """.format(table_surveyor=TABLE_SURVEYOR))
        con.commit()

        for row in cur:
            print('>> BAD SURVEYOR FULLNAME: "%s (%s)"' % (row[0], ' '.join(filter(lambda n: n is not None, row[1:]))))

        # Create and popluate a table mapping hollins fullname to surveyor fullname
        print('CREATE TABLE: {table_hollins_fullname} ...'.format(table_hollins_fullname=TABLE_HOLLINS_FULLNAME))

        cur.execute("""
            DROP TABLE IF EXISTS {table_hollins_fullname};
            CREATE TABLE {table_hollins_fullname}  (
              id serial PRIMARY KEY,
              hollins_fullname text NOT NULL UNIQUE,
              fullname text NOT NULL REFERENCES {table_surveyor} (fullname)
            );
        """.format(
            table_hollins_fullname=TABLE_HOLLINS_FULLNAME,
            table_hollins_map=TABLE_HOLLINS_MAP,
            table_surveyor=TABLE_SURVEYOR,
        ))

        cur.executemany("""
            INSERT INTO {table_hollins_fullname} (hollins_fullname, fullname)
            VALUES (%s, %s);
        """.format(table_hollins_fullname=TABLE_HOLLINS_FULLNAME), list(hollins_mapping.items()))
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))

        # Check every valid surveyor in hollins map has a mapping to the surveyor table
        cur.execute("""
            WITH q1 AS (
                SELECT id AS map_id,
                    regexp_split_to_table(upper(surveyor), '\s+&\s+') AS hollins_fullname
                FROM {table_hollins_map}
            )
            SELECT DISTINCT q1.hollins_fullname
            FROM q1
            LEFT JOIN {table_hollins_fullname} f USING (hollins_fullname)
            WHERE f.id IS NULL;
        """.format(
            table_hollins_map=TABLE_HOLLINS_MAP,
            table_hollins_fullname=TABLE_HOLLINS_FULLNAME
        ))
        con.commit()
        for row in cur:
            print('>> UNKNOWN SURVEYOR: "%s"' % row[0])

        # Create and popluate the signed_by table
        print('CREATE TABLE: {table_signed_by} ...'.format(table_signed_by=TABLE_SIGNED_BY))

        cur.execute("""
            DROP TABLE IF EXISTS {table_signed_by};
            CREATE TABLE {table_signed_by}  (
              map_id integer NOT NULL REFERENCES {table_map},
              surveyor_id integer NOT NULL REFERENCES {table_surveyor},
              PRIMARY KEY (map_id, surveyor_id)
            );
        """.format(
            table_signed_by=TABLE_SIGNED_BY,
            table_map=TABLE_MAP,
            table_surveyor=TABLE_SURVEYOR
        ))

        # Maps without surveyors in the surveyor list, i.e. UNKNOWN, etc.,
        # don't get a record in the signed by table.
        cur.execute("""
            WITH q1 AS (
                SELECT id AS map_id,
                    regexp_split_to_table(upper(surveyor), '\s*&\s*') AS hollins_fullname
                FROM {table_hollins_map}
            ), q2 AS (
                SELECT q1.map_id, s.id as surveyor_id
                FROM q1
                JOIN {table_hollins_fullname} f USING (hollins_fullname)
                JOIN {table_surveyor} s USING (fullname)
            )
            INSERT INTO {table_signed_by} (map_id, surveyor_id)
            SELECT map_id, surveyor_id from q2;
        """.format(
            table_hollins_map=TABLE_HOLLINS_MAP,
            table_hollins_fullname=TABLE_HOLLINS_FULLNAME,
            table_surveyor=TABLE_SURVEYOR,
            table_signed_by=TABLE_SIGNED_BY
        ))
        con.commit()

        print('INSERT (HOLLINS): %d rows affected.' % (cur.rowcount))

        # Read additional data from the map XLSX file
        ws = load_workbook(filename=XLSX_DATA_MAP, read_only=True).active
        HEADER_LINES = 1

        maps = []
        for map in ws.iter_rows(min_row=HEADER_LINES + 1):
            maps.append(tuple(c.value for c in map))

        cur.executemany("""
            WITH q1 AS (
            SELECT
                (%s)::text maptype,
                (%s)::integer book,
                (%s)::integer page,
                (%s)::integer npages,
                (%s)::date recdate,
                regexp_split_to_table(upper((%s)::text), '\s*&\s*') hollins_fullname,
                (%s)::text client,
                (%s)::text description,
                (%s)::text trs,
                (%s)::text note
            )
            INSERT INTO {table_signed_by} (map_id, surveyor_id)
            SELECT m.id map_id, s.id as surveyor_id
            FROM q1
            JOIN {table_maptype} t ON q1.maptype = t.maptype
            JOIN {table_map} m ON t.id = m.maptype_id
                AND q1.book = m.book AND q1.page = m.page
            JOIN {table_hollins_fullname} f ON q1.hollins_fullname = f.hollins_fullname
            JOIN {table_surveyor} s ON f.fullname = s.fullname
            ;
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE,
            table_hollins_fullname=TABLE_HOLLINS_FULLNAME,
            table_surveyor=TABLE_SURVEYOR,
            table_signed_by=TABLE_SIGNED_BY
        ), maps)
        con.commit()

        print('INSERT (EXTRAS): ' + str(cur.rowcount) + ' rows effected.')


        # Vacuum up dead tuples from the update
        tables = (
            TABLE_SURVEYOR,
            TABLE_SIGNED_BY
        )

        # Vacuum must run outside of a transaction
        con.autocommit = True
        for t in tables:
            cur.execute('VACUUM FREEZE ' + t)

if __name__ == '__main__':

    print('\nLoading tables ... ')
    startTime = time.time()

    load_surveyor()

    endTime = time.time()
    print('{0:.3f} sec'.format(endTime - startTime))
