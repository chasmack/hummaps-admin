import psycopg2
from openpyxl import load_workbook
import time

from const import *

# load_map.py - create and load the map and maptype tables

def load_map():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Load the maptype table
        print('CREATE TABLE: {table_maptype} ...'.format(table_maptype=TABLE_MAPTYPE))

        cur.execute("""
            DROP TABLE IF EXISTS {table_maptype} CASCADE;
            CREATE TABLE {table_maptype}  (
               id serial PRIMARY KEY,
               maptype text,
               abbrev text
            );
        """.format(table_maptype=TABLE_MAPTYPE))

        cur.executemany("""
            INSERT INTO {table_maptype} (maptype, abbrev) VALUES (%s, %s);
        """.format(table_maptype=TABLE_MAPTYPE), MAPTYPES)
        rowcount = cur.rowcount
        con.commit()

        # Check we have all the maptypes listed in hollins
        cur.execute("""
            SELECT DISTINCT m.id, maptype
            FROM {table_hollins_map} m
            LEFT JOIN {table_maptype} t USING (maptype)
            WHERE t.id is NULL;
        """.format(
            table_hollins_map=TABLE_HOLLINS_MAP,
            table_maptype=TABLE_MAPTYPE,
        ))
        con.commit()
        assert cur.rowcount == 0

        print('INSERT: ' + str(rowcount) + ' rows effected.')

        # Load the map tables from Hollins update
        print('CREATE TABLE: {table_map} ...'.format(table_map=TABLE_MAP))

        cur.execute("""
            DROP TABLE IF EXISTS {table_map} CASCADE;
            CREATE TABLE {table_map}  (
              id serial PRIMARY KEY,
              maptype_id integer REFERENCES {table_maptype},
              book integer,
              page integer,
              npages integer,
              recdate date,
              client text,
              description text,
              note text
            );
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE
        ))

        cur.execute("""
            INSERT INTO {table_map} (
               id, maptype_id, book, page, npages,
               recdate, client, description, note
            )
            SELECT
              {table_hollins_map}.id,
              {table_maptype}.id,
              {table_hollins_map}.book,
              {table_hollins_map}.firstpage,
              {table_hollins_map}.lastpage - {table_hollins_map}.firstpage + 1,
              {table_hollins_map}.recdate,
              {table_hollins_map}.donefor,
              {table_hollins_map}.descript,
              {table_hollins_map}.comment
            FROM {table_hollins_map}
            LEFT JOIN {table_maptype} USING (maptype)
            ;
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE,
            table_hollins_map=TABLE_HOLLINS_MAP
        ))
        con.commit()

        print('INSERT (HOLLINS): ' + str(cur.rowcount) + ' rows effected.')

        # Read additional map data from the XLSX file
        ws = load_workbook(filename=XLSX_DATA_MAP, read_only=True).active
        HEADER_LINES = 1

        maps = []
        for map in ws.iter_rows(min_row=HEADER_LINES + 1):
            maps.append(tuple(c.value for c in map))

        cur.executemany("""
            -- Reset the primary key sequence
            SELECT setval('{sequence_map_id}', max(id) + 1) FROM {table_map};
            WITH q1 AS (
            SELECT
                (%s)::text maptype,
                (%s)::integer book,
                (%s)::integer page,
                (%s)::integer npages,
                (%s)::date recdate,
                (%s)::text surveyor,
                (%s)::text client,
                (%s)::text description,
                (%s)::text trs,
                (%s)::text note
            )
            INSERT INTO {table_map} (
               maptype_id, book, page, npages,
               recdate, client, description, note
            )
            SELECT t.id, q1.book, q1.page, q1.npages,
                q1.recdate, q1.client, q1.description, q1.note
            FROM q1
            JOIN {table_maptype} t USING (maptype)
            ;
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE,
            sequence_map_id=SEQUENCE_MAP_ID
        ), maps)
        con.commit()

        print('INSERT (EXTRAS): ' + str(cur.rowcount) + ' rows effected.')

        # Read parcel map numbers from the XLSX file
        ws = load_workbook(filename=XLSX_DATA_PM, read_only=True).active
        HEADER_LINES = 1

        maps = []
        for map in ws.iter_rows(min_row=HEADER_LINES + 1):
            maps.append(tuple(c.value for c in map))

        cur.executemany("""
            WITH q1 AS (
              SELECT m.id map_id, t.maptype, m.book, m.page,
                regexp_replace(v.pm, '(\d+)(.*)', ' (PM\\1)\\2') pm
              FROM (
                VALUES (
                  (%s)::text, (%s)::integer, (%s)::integer, (%s)::text, (%s)::text
                )
              ) AS v (maptype, book, page, pm, note)
              LEFT JOIN {table_map} m ON v.book = m.book AND v.page = m.page
              LEFT JOIN {table_maptype} t ON m.maptype_id = t.id
              WHERE v.maptype = t.maptype
            )
            UPDATE {table_map} m
            SET client = client || q1.pm
            FROM q1
            WHERE m.id = q1.map_id
            ;
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE
        ), maps)
        con.commit()

        print('UPDATE (PARCEL NUMBER): ' + str(cur.rowcount) + ' rows effected.')

        # Read tract numbers from the XLSX file
        ws = load_workbook(filename=XLSX_DATA_TRACT, read_only=True).active
        HEADER_LINES = 1

        maps = []
        for map in ws.iter_rows(min_row=HEADER_LINES + 1):
            maps.append(tuple(c.value for c in map))

        cur.executemany("""
            WITH q1 AS (
              SELECT m.id map_id, t.maptype, m.book, m.page,
                regexp_replace(v.pm, '(\d+)(.*)', ' (TR\\1)\\2') pm
              FROM (
                VALUES (
                  (%s)::text, (%s)::integer, (%s)::integer, (%s)::text, (%s)::text
                )
              ) AS v (maptype, book, page, pm, note)
              LEFT JOIN {table_map} m ON v.book = m.book AND v.page = m.page
              LEFT JOIN {table_maptype} t ON m.maptype_id = t.id
              WHERE v.maptype = t.maptype
            )
            UPDATE {table_map} m
            SET client = client || q1.pm
            FROM q1
            WHERE m.id = q1.map_id
            ;
        """.format(
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE
        ), maps)
        con.commit()

        print('UPDATE (TRACT NUMBER): ' + str(cur.rowcount) + ' rows effected.')

        FULL_TOWNSHIP = ''.join((
            'S1,S2,S3,S4,S5,S6,',
            'S7,S8,S9,S10,S11,S12,',
            'S13,S14,S15,S16,S17,S18,',
            'S19,S20,S21,S22,S23,S24,',
            'S25,S26,S27,S28,S29,S30,',
            'S31,S32,S33,S34,S35,S36'
        ))

        cur.execute("""
            WITH q1 AS (
              SELECT m.id,
                regexp_replace(m.description, '{full_township}', 'FULL TOWNSHIP') description
              FROM {table_map} m
              WHERE m.description ~ '{full_township}'
            )
            UPDATE {table_map} m
            SET description = q1.description
            FROM q1
            WHERE m.id = q1.id
            ;
        """.format(
            table_map=TABLE_MAP,
            full_township=FULL_TOWNSHIP
        ))
        con.commit()

        print('UPDATE (MAP DESCRIPTION): ' + str(cur.rowcount) + ' rows effected.')

        # Vacuum up dead tuples from the update
        tables = (
            TABLE_MAP,
            TABLE_MAPTYPE
        )

        # Vacuum must run outside of a transaction
        con.autocommit = True
        for t in tables:
            cur.execute('VACUUM FREEZE ' + t)


if __name__ == '__main__':

    print('\nCreating staging tables ... ')
    startTime = time.time()

    load_map()

    endTime = time.time()
    print('{0:.3f} sec'.format(endTime - startTime))
