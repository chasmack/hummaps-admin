import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import psycopg2
from os.path import join
import time
import re

# init_update.py - create the hummaps_update schema and load data from hollins

UPDATE_DIR = 'data/update59'

PG_DATABASE = 'production'
PG_USER = 'ubuntu'
PG_HOST = 'localhost'
PG_PASSWORD = 'pg'

# IPV4 connection
# PG_DSN = 'dbname={database} user={user} host={host} password={password}'.format(
#     database=PG_DATABASE, user=PG_USER, host=PG_HOST, password=PG_PASSWORD
# )

# UNIX domain socket connection
PG_DSN = 'dbname={database} user={user}'.format(
    database=PG_DATABASE, user=PG_USER
)

SCHEMA_STAGING = 'hummaps_staging'
SCHEMA_UPDATE = 'hummaps_update'
TABLE_HOLLINS_MAP = SCHEMA_UPDATE + '.' + 'hollins_map'
TABLE_HOLLINS_MAP_QQ = SCHEMA_UPDATE + '.' + 'hollins_map_qq'
TABLE_HOLLINS_SUBSECTION_LIST = SCHEMA_UPDATE + '.' + 'hollins_subsection_list'
TABLE_HOLLINS_SURVEYOR = SCHEMA_UPDATE + '.' + 'hollins_surveyor'
TABLE_HOLLINS_TRS = SCHEMA_UPDATE + '.' + 'hollins_trs'
TABLE_SURVEYOR = SCHEMA_UPDATE + '.' + 'surveyor'
TABLE_CC = SCHEMA_UPDATE + '.' + 'cc'

XLSX_DATA_SURVEYOR = 'data/surveyor.xlsx'
XLSX_DATA_CC = 'data/cc.xlsx'


def init_staging():
    # Create the hummaps_staging schema

    print('CREATE SCHEMA: {schema} ...'.format(schema=SCHEMA_STAGING))

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        pg_qstr = """
            DROP SCHEMA IF EXISTS {schema} CASCADE;
            CREATE SCHEMA {schema};
        """.format(schema=SCHEMA_STAGING)
        cur.execute(pg_qstr)
        con.commit()


def init_update():
    # Create the hummaps_update schema

    print('CREATE SCHEMA: {schema} ...'.format(schema=SCHEMA_UPDATE))

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:
        pg_qstr = """
                DROP SCHEMA IF EXISTS {schema} CASCADE;
                CREATE SCHEMA {schema};
            """.format(schema=SCHEMA_UPDATE)
        cur.execute(pg_qstr)
        con.commit()


def load_hollins():
    # Load the hummaps_staging tables from XML data

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        print('CREATE TABLE: {table} ...'.format(table=TABLE_HOLLINS_SURVEYOR))

        pg_qstr = """
            DROP TABLE IF EXISTS {table};
            CREATE TABLE {table} (
              id serial PRIMARY KEY,
              fullname text,
              lastname text
            );
        """.format(table=TABLE_HOLLINS_SURVEYOR)
        cur.execute(pg_qstr)

        tree = ET.parse(join(UPDATE_DIR, 'surveyor.xml'))
        root = tree.getroot()
        rows = []
        keys = ('Surveyor', 'lastname')
        for rec in root:
            rows.append(list(rec.find(k).text for k in keys))

        pg_qstr = """
          INSERT INTO {table} (fullname, lastname) VALUES ( %s, %s );
        """.format(table=TABLE_HOLLINS_SURVEYOR)
        cur.executemany(pg_qstr, rows)
        con.commit()

        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Transfer contents of SubSectionList table to pg staging

        print('CREATE TABLE: {table} ...'.format(table=TABLE_HOLLINS_SUBSECTION_LIST))

        pg_qstr = """
            DROP TABLE IF EXISTS {table};
            CREATE TABLE {table} (
              id serial PRIMARY KEY,
              order_code integer,
              subsection text
            );
        """.format(table=TABLE_HOLLINS_SUBSECTION_LIST)
        cur.execute(pg_qstr)

        tree = ET.parse(join(UPDATE_DIR, 'subsectionlist.xml'))
        root = tree.getroot()
        rows = []
        keys = ('OrderCode', 'subsection')
        for rec in root:
            rows.append(list(rec.find(k).text for k in keys))

        pg_qstr = """
          INSERT INTO {table} (order_code, subsection) VALUES ( %s, %s );
        """.format(table=TABLE_HOLLINS_SUBSECTION_LIST)
        cur.executemany(pg_qstr, rows)
        con.commit()

        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Transfer contents of map table to pg staging
        print('CREATE TABLE: {table} ...'.format(table=TABLE_HOLLINS_MAP))

        pg_qstr = """
            DROP TABLE IF EXISTS {table};
            CREATE TABLE {table} (
              id integer PRIMARY KEY,
              maptype text,
              book integer,
              firstpage integer,
              lastpage integer,
              recdate date,
              surveyor text,
              donefor text,
              descript text,
              image text,
              comment text
            );
        """.format(table=TABLE_HOLLINS_MAP)
        cur.execute(pg_qstr)


        tree = ET.parse(join(UPDATE_DIR, 'map.xml'))
        root = tree.getroot()
        rows = []
        keys = ('ID', 'maptype', 'BOOK', 'FIRSTPAGE', 'LASTPAGE', 'RECDATE',
                'SURVEYOR', 'DONEFOR', 'DESCRIP', 'Picture', 'Comment')
        for rec in root:
            rows.append(list(c if c is None else c.text for c in (rec.find(k) for k in keys)))

        # Changed a couple field names from the original
        pg_qstr = """
            INSERT INTO {table} (
              id, maptype, book, firstpage, lastpage, recdate,
              surveyor, donefor, descript, image, comment
            ) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """.format(table=TABLE_HOLLINS_MAP)

        cur.executemany(pg_qstr, rows)
        con.commit()

        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Transfer contents of trs table to pg staging
        print('CREATE TABLE: {table} ...'.format(table=TABLE_HOLLINS_TRS))

        pg_qstr = """
            DROP TABLE IF EXISTS {table};
            CREATE TABLE {table} (
              id serial PRIMARY KEY,
              map_id integer,
              township text,
              range text,
              section text
            );
        """.format(table=TABLE_HOLLINS_TRS)
        cur.execute(pg_qstr)

        tree = ET.parse(join(UPDATE_DIR, 'trs.xml'))
        root = tree.getroot()
        rows = []
        keys = ('ID', 'TOWNSHIP', 'RANGE', 'SECTION')
        for rec in root:
            rows.append(list(rec.find(k).text for k in keys))

        pg_qstr = """
            INSERT INTO {table} (map_id, township, range, section)
            VALUES ( %s, %s, %s, %s );
        """.format(table=TABLE_HOLLINS_TRS)

        cur.executemany(pg_qstr, rows)
        con.commit()

        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')


def load_hollins_qq():
    # Split out the quarter-quarter section info from hollins_map into its own table

    print('CREATE TABLE: {table} ...'.format(table=TABLE_HOLLINS_MAP_QQ))

    # get a list of the qq section columns from map
    tree = ET.parse(join(UPDATE_DIR, 'map.xml'))
    root = tree.getroot()
    skip_tags = (
        'id', 'maptype', 'book', 'firstpage', 'lastpage',
        'recdate', 'surveyor', 'donefor', 'descrip', 'picture', 'comment')

    qq_cols = set()
    for row in root.findall('map'):
        for col in row:
            tag = col.tag.lower()
            if tag in skip_tags:
                continue
            qq_cols.add(tag[7:])
    qq_cols = sorted(qq_cols)

    with psycopg2.connect(PG_DSN) as con, con.cursor() as curs:

        # Create the maps_qq table
        pg_qstr = """
            DROP TABLE IF EXISTS {table_map_qq};
            CREATE TABLE {table_map_qq} (
              id integer PRIMARY KEY,
              FOREIGN KEY (id) REFERENCES {table_map}
            );
        """.format(table_map_qq=TABLE_HOLLINS_MAP_QQ, table_map=TABLE_HOLLINS_MAP)
        curs.execute(pg_qstr)
        for col in qq_cols:
            curs.execute("""
                ALTER TABLE {table_map_qq} ADD COLUMN "{col}" text;
            """.format(table_map_qq=TABLE_HOLLINS_MAP_QQ, col=col))
        con.commit()

        # Populate map_qq with maps with qq information
        for row in root.findall('map'):
            qq_cols = []
            for col in row:
                tag = col.tag.lower()
                if tag in skip_tags:
                    continue
                qq_cols.append((tag[7:], col.text))
            if len(qq_cols) == 0:
                continue
            id = row.find('ID').text
            qq_values = ', '.join("'%s'" % qq[1] for qq in qq_cols)
            qq_cols = ', '.join('"%s"' % qq[0] for qq in qq_cols)
            pg_qstr = """
                INSERT INTO {table_map_qq} (id, {qq_cols})
                VALUES ({id}, {qq_values});
            """.format(table_map_qq=TABLE_HOLLINS_MAP_QQ, id=id, qq_cols=qq_cols, qq_values=qq_values)
            curs.execute(pg_qstr)

        curs.execute('SELECT count(*) FROM {table_map_qq}'.format(table_map_qq=TABLE_HOLLINS_MAP_QQ))
        con.commit()
        print('INSERT: %d rows affected.' % curs.fetchone())


def load_surveyor():
    # Load the surveyor table from the Excel spreadsheet

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        print('CREATE TABLE: {table} ...'.format(table=TABLE_SURVEYOR))

        pg_qstr = """
            DROP TABLE IF EXISTS {table_surveyor};
            CREATE TABLE {table_surveyor} (
                id serial PRIMARY KEY,
                hollins_fullname text,
                fullname text,
                firstname text,
                secondname text,
                thirdname text,
                lastname text,
                suffix text,
                pls text,
                rce text
            );
        """.format(table_surveyor=TABLE_SURVEYOR)
        cur.execute(pg_qstr)
        con.commit()

        ws = load_workbook(filename=XLSX_DATA_SURVEYOR, read_only=True).active

        # skip over the headers
        values = []
        for row in ws.iter_rows(min_row=2):
            values.append(list(c.value for c in row))

        pg_qstr = """
            INSERT INTO {table_surveyor} (
              hollins_fullname, fullname, firstname, secondname, thirdname, lastname, suffix, pls, rce
            ) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """.format(table_surveyor=TABLE_SURVEYOR)
        cur.executemany(pg_qstr, values)
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))

def load_cc():
    # Load the cc table from the Excel spreadsheet

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        print('CREATE TABLE: {table_cc} ...'.format(table_cc=TABLE_CC))

        pg_qstr = """
           DROP TABLE IF EXISTS {table_cc};
           CREATE TABLE {table_cc} (
               id serial PRIMARY KEY,
               maptype text,
               book integer,
               firstpage integer,
               lastpage integer,
               recdate date,
               surveyor text,
               donefor text,
               doc_number text,
               pages integer
           );
       """.format(table_cc=TABLE_CC)
        cur.execute(pg_qstr)
        con.commit()

        ws = load_workbook(filename=XLSX_DATA_CC, read_only=True).active

        # skip over the headers
        values = []
        for row in ws.iter_rows(min_row=2):
            values.append(list(c.value for c in row))

        pg_qstr = """
                    INSERT INTO {table_cc} (
                      maptype, book, firstpage, lastpage, recdate, surveyor, donefor, doc_number, pages
                    ) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s);
                """.format(table_cc=TABLE_CC)
        cur.executemany(pg_qstr, values)
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))


def cleanup_update():
    # Vacuum up dead tuples from the update

    tables = (
        TABLE_HOLLINS_MAP,
        TABLE_HOLLINS_MAP_QQ,
        TABLE_HOLLINS_SUBSECTION_LIST,
        TABLE_HOLLINS_SURVEYOR,
        TABLE_HOLLINS_TRS,
        TABLE_SURVEYOR,
        TABLE_CC,
    )

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        # Vacuum must run outside of a transaction
        con.autocommit = True
        for t in tables:
            cur.execute('VACUUM FREEZE ' + t)


if __name__ == '__main__':

    print('\nLoading tables ... ')
    startTime = time.time()

    init_update()
    load_hollins()
    load_hollins_qq()
    load_surveyor()
    load_cc()
    cleanup_update()

    init_staging()

    endTime = time.time()
    print('{0:.3f} sec'.format(endTime - startTime))
