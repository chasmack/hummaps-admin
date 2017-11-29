from openpyxl import load_workbook
import psycopg2
import boto3
import time
import re

from const import *

# load_cc.py - create and load the cc and cc_image tables from XLSX data

def load_cc():
    # Load the cc table from cc.xlsx.

    with psycopg2.connect(PG_DSN) as con, con.cursor() as cur:

        # Create the cc table
        print('CREATE TABLE: {table_cc} ...'.format(table_cc=TABLE_CC))

        cur.execute("""
            DROP TABLE IF EXISTS {table_cc} CASCADE;
            CREATE TABLE {table_cc} (
              id serial PRIMARY KEY,
              map_id integer REFERENCES {table_map},
              doc_number text,
              npages integer
            );
        """.format(
            table_cc=TABLE_CC,
            table_map=TABLE_MAP
        ))
        con.commit()

        # Read the XLSX data
        ws = load_workbook(filename=XLSX_DATA_CC, read_only=True).active
        HEADER_LINES = 1

        cc_data = []
        for row in ws.iter_rows(min_row=HEADER_LINES + 1):
            cc_data.append(list(c.value for c in row))

        cur.executemany("""
            WITH q1 AS (
            SELECT
                (%s)::text maptype,
                (%s)::integer book,
                (%s)::integer page,
                (%s)::date recdate,
                (%s)::text surveyor,
                (%s)::text doc_number,
                (%s)::integer npages
            )
            INSERT INTO {table_cc} (map_id, doc_number, npages)
            SELECT map_id, doc_number, npages
            FROM q1
            LEFT JOIN (
              SELECT m.id map_id, t.maptype, m.book, m.page, m.recdate
              FROM {table_map} m
              JOIN {table_maptype} t
              ON m.maptype_id = t.id
            ) AS s1
            ON q1.maptype = s1.maptype
            AND q1.book = s1.book
            AND q1.page = s1.page
            AND q1.recdate = s1.recdate
            ;
        """.format(
            table_cc=TABLE_CC,
            table_map=TABLE_MAP,
            table_maptype=TABLE_MAPTYPE
        ), cc_data)
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))

        # Create the cc_image table
        print('CREATE TABLE: {table_cc_image} ...'.format(table_cc_image=TABLE_CC_IMAGE))

        cur.execute("""
            DROP TABLE IF EXISTS {table_cc_image};
            CREATE TABLE {table_cc_image} (
              id serial PRIMARY KEY,
              cc_id integer REFERENCES {table_cc},
              page integer,
              imagefile text
            );
        """.format(
            table_cc_image=TABLE_CC_IMAGE,
            table_cc=TABLE_CC
        ))
        con.commit()

        # Get a list of cc images from the maps bucket
        s3 = boto3.resource('s3')
        bucket = s3.Bucket(S3_BUCKET_MAPS)

        # Put each imagefile in a length = 1 list to be later extended
        imagefiles = list([obj.key] for obj in bucket.objects.filter(Prefix='map/cc/'))

        # Need the number of pages in a document for new style document number
        max_pages = {}
        for f in imagefiles:
            imagefile = f[0]
            basename, page = re.match('.*/(.*)-(\d{3})', imagefile).groups()
            if basename not in max_pages:
                max_pages[basename] = int(page)
            else:
                max_pages[basename] = max(int(page), max_pages[basename])

        # Add the doc number and page to each imagefile
        for f in imagefiles:
            imagefile = f[0]
            filename = imagefile.split('/')[-1]
            basename = re.match('(.*)-', filename).group(1)
            s1, cc_type, s2, page = (p.lstrip('0') for p in filename.rstrip('.jpg').split('-'))
            if cc_type == 'or':
                doc_number = '%s OR %s' % (s1, s2)
            else:
                doc_number = '%s-%s-%d' % (s1, s2, max_pages[basename])
            f[0:3] = [doc_number, int(page), imagefile]

        # Insert imagefile records into the cc image table
        cur.executemany("""
            WITH q1 AS (
                SELECT
                  (%s)::text doc_number,
                  (%s)::integer page,
                  (%s)::text imagefile
            ), q2 AS (
                SELECT cc.id cc_id, page, imagefile
                FROM q1
                LEFT JOIN {table_cc} cc USING (doc_number)
            )
            INSERT INTO {table_cc_image} (cc_id, page, imagefile)
            SELECT cc_id, page, imagefile
            FROM q2
            ;
        """.format(
            table_cc=TABLE_CC,
            table_cc_image=TABLE_CC_IMAGE
        ), imagefiles)
        con.commit()

        print('INSERT: %d rows affected.' % (cur.rowcount))

        # Vacuum must run outside of a transaction
        con.autocommit = True
        for table in (TABLE_CC, TABLE_CC_IMAGE):
            cur.execute('VACUUM FREEZE ' + table)

if __name__ == '__main__':

    print('\nLoading tables ... ')
    startTime = time.time()

    load_cc()

    endTime = time.time()
    print('{0:.3f} sec'.format(endTime - startTime))
